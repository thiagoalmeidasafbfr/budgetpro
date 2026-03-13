import io
from decimal import Decimal, InvalidOperation
from typing import List, Optional
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.models.models import Budget, CostCenter, Account, AuditLog, AuditAction

MONTH_COLS = ["jan", "feb", "mar", "apr", "may", "jun",
              "jul", "aug", "sep", "oct", "nov", "dec"]

IMPORT_COLUMNS = [
    "CostCenterCode", "AccountCode",
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Notes",
]


def import_budget_excel(
    db: Session,
    version_id: UUID,
    file_content: bytes,
    user_id: UUID,
) -> dict:
    """Import budget data from an Excel file.

    Expected columns: CostCenterCode, AccountCode, Jan..Dec, Notes
    """
    errors: List[str] = []
    imported = 0

    try:
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        return {"imported": 0, "errors": [f"Failed to read Excel file: {str(e)}"]}

    ws = wb.active
    if ws is None:
        return {"imported": 0, "errors": ["Excel file has no active worksheet"]}

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"imported": 0, "errors": ["Excel file has no data rows (only header or empty)"]}

    # Parse header row to find column indices
    header = [str(cell).strip() if cell else "" for cell in rows[0]]
    col_map = {}
    for expected_col in IMPORT_COLUMNS:
        for idx, h in enumerate(header):
            if h.lower() == expected_col.lower():
                col_map[expected_col] = idx
                break

    # Validate required columns exist
    missing_cols = []
    for col in ["CostCenterCode", "AccountCode"]:
        if col not in col_map:
            missing_cols.append(col)
    if missing_cols:
        return {"imported": 0, "errors": [f"Missing required columns: {', '.join(missing_cols)}"]}

    # Cache cost centers and accounts by code
    cost_centers = {cc.code: cc for cc in db.query(CostCenter).all()}
    accounts = {acc.code: acc for acc in db.query(Account).all()}

    for row_num, row in enumerate(rows[1:], start=2):
        cc_code_val = row[col_map["CostCenterCode"]] if col_map.get("CostCenterCode") is not None else None
        acc_code_val = row[col_map["AccountCode"]] if col_map.get("AccountCode") is not None else None

        if not cc_code_val and not acc_code_val:
            continue  # Skip empty rows

        cc_code = str(cc_code_val).strip() if cc_code_val else ""
        acc_code = str(acc_code_val).strip() if acc_code_val else ""

        if not cc_code:
            errors.append(f"Row {row_num}: Missing CostCenterCode")
            continue
        if not acc_code:
            errors.append(f"Row {row_num}: Missing AccountCode")
            continue

        cost_center = cost_centers.get(cc_code)
        if not cost_center:
            errors.append(f"Row {row_num}: Cost center '{cc_code}' not found")
            continue

        account = accounts.get(acc_code)
        if not account:
            errors.append(f"Row {row_num}: Account '{acc_code}' not found")
            continue

        # Parse month values
        month_values = {}
        has_error = False
        for month_name, col_attr in zip(
            ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"],
            MONTH_COLS,
        ):
            raw_val = row[col_map[month_name]] if col_map.get(month_name) is not None and col_map[month_name] < len(row) else None
            if raw_val is None or str(raw_val).strip() == "":
                month_values[col_attr] = Decimal("0")
            else:
                try:
                    month_values[col_attr] = Decimal(str(raw_val).replace(",", "."))
                except (InvalidOperation, ValueError):
                    errors.append(f"Row {row_num}: Invalid value '{raw_val}' for {month_name}")
                    has_error = True
                    break

        if has_error:
            continue

        # Parse notes
        notes_val = row[col_map["Notes"]] if col_map.get("Notes") is not None and col_map["Notes"] < len(row) else None
        notes = str(notes_val).strip() if notes_val else None

        # Find or create budget row
        existing = db.query(Budget).filter(
            Budget.version_id == version_id,
            Budget.cost_center_id == cost_center.id,
            Budget.account_id == account.id,
        ).first()

        if existing:
            for col_attr, value in month_values.items():
                setattr(existing, col_attr, value)
            if notes is not None:
                existing.notes = notes
        else:
            budget = Budget(
                version_id=version_id,
                cost_center_id=cost_center.id,
                account_id=account.id,
                notes=notes,
                **month_values,
            )
            db.add(budget)

        imported += 1

    if imported > 0:
        # Log the import action
        audit = AuditLog(
            user_id=user_id,
            entity_type="Budget",
            entity_id=version_id,
            action=AuditAction.create,
            changes={"action": "excel_import", "rows_imported": imported, "errors_count": len(errors)},
        )
        db.add(audit)
        db.commit()

    wb.close()
    return {"imported": imported, "errors": errors}


def export_budget_excel(
    db: Session,
    version_id: UUID,
    version_name: str,
) -> io.BytesIO:
    """Export budget data to a styled Excel workbook."""
    budgets = (
        db.query(Budget)
        .filter(Budget.version_id == version_id)
        .options(
            joinedload(Budget.cost_center),
            joinedload(Budget.account),
        )
        .order_by(Budget.cost_center_id, Budget.account_id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = version_name[:31] if len(version_name) > 31 else version_name

    # Define styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_format = '#,##0.00'

    # Headers
    headers = [
        "Centro de Custo", "Codigo CC", "Conta", "Codigo Conta", "Categoria",
        "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
        "Jul", "Ago", "Set", "Out", "Nov", "Dez",
        "Total", "Notas",
    ]

    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Category translation
    category_labels = {
        "revenue": "Receita",
        "cost": "Custo",
        "expense": "Despesa",
        "investment": "Investimento",
    }

    # Data rows
    for row_idx, budget in enumerate(budgets, start=2):
        cc_name = budget.cost_center.name if budget.cost_center else ""
        cc_code = budget.cost_center.code if budget.cost_center else ""
        acc_name = budget.account.name if budget.account else ""
        acc_code = budget.account.code if budget.account else ""
        acc_category = category_labels.get(budget.account.category.value, "") if budget.account else ""

        row_data = [
            cc_name, cc_code, acc_name, acc_code, acc_category,
            float(budget.jan or 0), float(budget.feb or 0), float(budget.mar or 0),
            float(budget.apr or 0), float(budget.may or 0), float(budget.jun or 0),
            float(budget.jul or 0), float(budget.aug or 0), float(budget.sep or 0),
            float(budget.oct or 0), float(budget.nov or 0), float(budget.dec or 0),
            float(budget.total or 0),
            budget.notes or "",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Apply number format to monetary columns (columns 6-18: Jan-Dec + Total)
            if 6 <= col_idx <= 18:
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right")

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(budgets) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                cell_len = len(str(cell_value))
                if cell_len > max_length:
                    max_length = cell_len
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Write to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    return output
